import openpyxl
from openpyxl.styles import Font, Alignment, numbers, Border, Side
import os
from commons.log_config import FILE_PATH

# Crear un nuevo libro de trabajo y una hoja de cálculo
wb = openpyxl.Workbook()
ws = wb.active

# Lista de datos a escribir en la hoja

datos = [
    ['Nombre', 'Edad', 'Ciudad'],
    ['Lucas', 22, 'Tucumán'],
    ['Ana', 30, 'Buenos Aires'],
    ['Carlos', 25, 'Córdoba'],
    ['Juan', 32, 'Rio Negro']
]

ws.append(['hola'])

# Agregar los datos a la hoja de cálculo
for indice, fila in enumerate(datos[2:4]):
    print(indice)
    ws.append(fila)

# for celda in ws[2]:
#     print(celda.value)

# for celda in ws['C']:
#     print(celda.value)

for fila in ws:
    for celda in fila:
        celda.alignment = Alignment(horizontal="center", vertical="center")

# Guardar el libro de trabajo en el directorio del script
nombre_archivo = 'datos_excel.xlsx'
wb.save(os.path.join(FILE_PATH, nombre_archivo))

print(f'El archivo {nombre_archivo} ha sido creado y guardado en el directorio del script.')
